#!/usr/bin/env python
# coding: utf-8

# In[6]:


import requests
from bs4 import BeautifulSoup

# USDC官方網站的URL
url = "https://www.circle.com/en/usdc"

# 建立一個session物件，以便在多個頁面之間共用cookie
session = requests.Session()

# 抓取網頁內容
response = session.get(url)

# 解析網頁內容
soup = BeautifulSoup(response.content, 'html.parser')

reserve = soup.find('div', {'class': 'chartjs-wrapper usdc_chart_data'})

# 找到包含數值的canvas標籤
canvas = soup.find('canvas', {'id': 'usdc_chartjs_canvas'})

date = soup.find('div', {'class': 'col pb-2'})

# 找到包含日期的<h6>標籤
date_tag = date.find("h6")

# 提取日期內容
date = date_tag.get_text(strip=True)

# 從canvas標籤的data-usdc-cash屬性中提取第一個數值
cash_value = canvas['data-usdc-cash']

# 從canvas標籤的data-usdc-in-circulation屬性中提取第二個數值
shortdated_value = canvas['data-usdc-in-circulation']

# 將字符串轉換成浮點數
cash_value = float(cash_value)
shortdated_value = float(shortdated_value)


# 檢查爬取資料 
print({'cash_value': cash_value, 'shortdated_value' : shortdated_value, 'date' : date})



###---匯入excel---###
from openpyxl import workbook, load_workbook 
from openpyxl.utils.dataframe import dataframe_to_rows

import openpyxl

# 打開Excel檔案
workbook = openpyxl.load_workbook('/Users/benson/Desktop/助理stuff/stablecoin/crawling.xlsx')

# 獲取要寫入的工作表
ws = workbook['data']

# 在工作表中添加標題行
# ws.append(['Cash', 'shoerdated_value', 'Date'])


# 確認最後一個有數值的單元格是否等於date
last_value = None
for row in ws.iter_rows(min_row=1, max_col=3, max_row=ws.max_row):
    cell_value = row[2].value
    if cell_value is not None:
        last_value = cell_value

if last_value != date:
    # 如果最後一個有數值的單元格不等於date，則將資料加入檔案中
    ws.append([cash_value, shortdated_value, date])
    workbook.save('/Users/benson/Desktop/助理stuff/stablecoin/crawling.xlsx')
else:
    print('already added')


# ## 自動運行封包

# In[4]:


# import datetime
# import schedule
# import time
# import requests
# from bs4 import BeautifulSoup
# from openpyxl import workbook, load_workbook 
# from openpyxl.utils.dataframe import dataframe_to_rows
# import openpyxl


# def job():
#     # 在這裡放置你要自動運行的程式碼
#     url = "https://www.circle.com/en/usdc"
#     session = requests.Session()
#     response = session.get(url)
#     soup = BeautifulSoup(response.content, 'html.parser')
#     reserve = soup.find('div', {'class': 'chartjs-wrapper usdc_chart_data'})
#     canvas = soup.find('canvas', {'id': 'usdc_chartjs_canvas'})
#     date = soup.find('div', {'class': 'col pb-2'})
#     date_tag = date.find("h6")
#     date = date_tag.get_text(strip=True)
#     cash_value = canvas['data-usdc-cash']
#     shortdated_value = canvas['data-usdc-in-circulation']
#     cash_value = float(cash_value)
#     shortdated_value = float(shortdated_value)
    
#     workbook = openpyxl.load_workbook('/Users/benson/Desktop/助理stuff/stablecoin/crawling.xlsx')
#     ws = workbook['data']
#     ws.append(['Cash', 'shoerdated_value', 'Date'])
#     ws.append([cash_value, shortdated_value, date])
#     workbook.save('/Users/benson/Desktop/助理stuff/stablecoin/crawling.xlsx')   
#     print("程式已運行！")


# In[3]:


# import schedule
# import time

# schedule.every().day.at("10:30").do(job) # 設定每天10:30自動執行job函數

# while True:
#     schedule.run_pending()
#     time.sleep(10)

